from django.shortcuts import render
from django.views.generic import ListView
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Proyecto,Empresa
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .forms import AgregarUsuarioForm,LoginForm
from django.core.exceptions import PermissionDenied

from django.http import JsonResponse, HttpResponse
import json
# Create your views here.
def home(request):
    return render(request, 'construccion1app/home.html')


from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView
from .models import Proyecto

class MisProyectosViewHome(LoginRequiredMixin, ListView):
    template_name = 'construccion1app/mis_proyectosHome.html'
    context_object_name = 'proyectos'

    def get_queryset(self):
        user = self.request.user
        empresa_buscar = self.request.GET.get('empresa_buscar', '').strip()

        if user.is_superuser:
            # Superusuario ve todos los proyectos o filtra por nombre de empresa
            queryset = Proyecto.objects.all()
            
            if empresa_buscar:
                # Búsqueda por coincidencias en el nombre de la empresa (case-insensitive)
                queryset = queryset.filter(empresa__nombre__icontains=empresa_buscar)
            
            return queryset

        if user.tipo_usuario in ['superadmin_empresa', 'admin_empresa'] and user.empresa:
            # Administradores de empresa ven todos los proyectos de su empresa
            return Proyecto.objects.filter(empresa=user.empresa)

        if user.empresa:
            # Usuarios normales solo ven proyectos a los que tienen acceso
            return Proyecto.objects.filter(
                empresa=user.empresa,
                usuarios_acceso=user
            )

        # Si no tiene empresa ni proyectos asignados
        return Proyecto.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Solo agregar lista de empresas si es superusuario
        if self.request.user.is_superuser:
            from .models import Empresa
            context['empresas'] = Empresa.objects.all().order_by('nombre')
        
        # Agregar avance_total como atributo temporal a cada proyecto
        proyectos = context['proyectos']
        for proyecto in proyectos:
            proyecto.avance_total = proyecto.calcular_avance()
        
        # Pasar los objetos directamente (no diccionarios)
        context['proyectos_data'] = proyectos
        
        return context

@login_required
def editar_visibilidad(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    # Solo superadmin_empresa o admin_empresa pueden acceder
    if not request.user.tipo_usuario in ['superadmin_empresa', 'admin_empresa']:
        return redirect('mis_proyectosHome')  # O mostrar un mensaje de error

    # Obtenemos los usuarios de tipo supervisor_constructor o calidad
    usuarios = Usuario.objects.filter(tipo_usuario__in=['supervisor_constructor', 'calidad'], empresa=request.user.empresa)
    
    if request.method == 'POST':
        seleccionados_ids = request.POST.getlist('usuarios')
        proyecto.usuarios_acceso.set(seleccionados_ids)
        return redirect('mis_proyectos')

    context = {
        'proyecto': proyecto,
        'usuarios': usuarios,
    }
    return render(request, 'construccion1app/editar_visibilidad.html', context)

    
@csrf_exempt
def login_view(request):
    if request.method == "POST":
        form = LoginForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('mis_proyectos')  # Cambia 'home' por tu URL de inicio
        else:
            messages.error(request, "Usuario o contraseña incorrectos")
    else:
        form = LoginForm()
    return render(request, 'construccion1app/login.html', {"form" :form})


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

# construccion1app/views.py
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Usuario

# views.py
@login_required
def agregarusuario(request):
    # Solo SuperAdministrador de Empresa y Administrador de Empresa
    if request.user.tipo_usuario not in ['superadmin_empresa', 'admin_empresa']:
        messages.error(request, "No tienes permisos para agregar usuarios.")
        return redirect('mis_proyectos')

    if request.method == "POST":
        form = AgregarUsuarioForm(request.POST, creador=request.user)
        if form.is_valid():
            nuevo_usuario = form.save(commit=False)
            nuevo_usuario.empresa = request.user.empresa  # Empresa automática
            nuevo_usuario.save()
            messages.success(request, f"Usuario {nuevo_usuario.username} creado correctamente.")
            return redirect('mis_proyectos')
    else:
        form = AgregarUsuarioForm(creador=request.user)

    return render(request, 'construccion1app/agregar_usuario.html', {'form': form})

@login_required
def lista_usuarios_empresa(request):
    # Solo SuperAdminEmpresa y AdminEmpresa pueden acceder
    if request.user.tipo_usuario not in ["superadmin_empresa", "admin_empresa"]:
        messages.error(request, "No tienes permisos para ver esta lista.")
        return redirect('mis_proyectos')

    # Filtrar usuarios de la misma empresa
    usuarios = Usuario.objects.filter(empresa=request.user.empresa)

    return render(request, 'construccion1app/lista_usuarios.html', {'usuarios': usuarios})


@login_required
def lista_empresas(request):
    # Superusuario no puede acceder
   

    # Mostrar solo la empresa a la que pertenece el usuario
    if request.user.empresa:
        empresas = Empresa.objects.filter(id=request.user.empresa.id)
    else:
        empresas = Empresa.objects.none()  # No tiene empresa

    return render(request, 'construccion1app/lista_empresas.html', {'empresas': empresas})



@login_required
def mi_perfil(request):
    # Bloqueo para superuser


    usuario = request.user  # Usuario logueado
    empresa = getattr(usuario, 'empresa', None)  # Solo si tiene empresa asignada

    context = {
        'usuario': usuario,
        'empresa': empresa,
    }
    return render(request, 'construccion1app/miperfil.html', context)


from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render
from decimal import Decimal, ROUND_HALF_UP

@login_required
def dashboard_proyecto(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    # Validar si el usuario puede ver el formulario
    mostrar_formulario = request.user.tipo_usuario in ["superadmin_empresa", "admin_empresa"]

    # Manejo del formulario
    if mostrar_formulario and request.method == "POST":
        form = NivelForm(request.POST, proyecto=proyecto)
        if form.is_valid():
            nivel = form.save(commit=False)
            nivel.proyecto = proyecto
            nivel.save()
            messages.success(request, "Nivel creado exitosamente.")
            form = NivelForm(proyecto=proyecto)  # limpiar formulario después de guardar
        else:
            messages.error(request, "Error al crear el nivel. Verifica los datos.")
    else:
        form = NivelForm(proyecto=proyecto) if mostrar_formulario else None

    # Construcción de datos de niveles y espacios
    datos_niveles = []
    for nivel in proyecto.niveles.all():
        datos_espacios = []
        for espacio in nivel.espacios.all():
            actividades = [
                {
                    "id": a.id,
                    "nombre": a.nombre,
                    "avance": a.avance,
                    "incidencia": a.incidencia,
                    "ponderado": a.calcular_avance_ponderado(),
                    "asignado": a.asignado,
                    "estado_ejecucion": a.get_estado_ejecucion_display(),
                    "habilitada": a.habilitada,
                    "predecesora": a.predecesora.nombre if a.predecesora else None,
                    "sucesora": a.sucesora.nombre if a.sucesora else None,               
                    "fecha_inicio": a.fecha_inicio,
                    "fecha_fin": a.fecha_fin,
                    "estado_asignacion": a.estado_asignacion,
                    "archivo_informacion": a.archivo_informacion,  # <-- aquí
                    "justificacion": a.justificacion,  # <-- aquí

                    
                }
                for a in espacio.actividades.all().order_by('id')  # <-- aquí
            ]

            # 👉 calcular incidencias de actividades de este espacio
            total_incidencia = espacio.actividades.aggregate(total=Sum("incidencia"))["total"] or 0
            total_incidencia = float(round(total_incidencia, 2))
            print(f"Espacio '{espacio.nombre}' tiene total de incidencias: {total_incidencia}")

            datos_espacios.append({
                "id": espacio.id,
                "nombre": espacio.nombre,
                "avance": sum(a["ponderado"] for a in actividades),
                "incidencia": espacio.incidencia,
                "ponderado": espacio.calcular_avance_ponderado(),
                "total_incidencia": total_incidencia,  # 👈 cada espacio tiene su propio total
                "actividades": actividades,
            })

        datos_niveles.append({
            "id": nivel.id,
            "nombre": nivel.nombre,
            "avance": sum(e["ponderado"] for e in datos_espacios),
            "incidencia": nivel.incidencia,
            "ponderado": nivel.calcular_avance_ponderado(),
            "espacios": datos_espacios,
        })

    avance_total = proyecto.calcular_avance()

    # Renderizar el dashboard con formulario y niveles
    # ❌ QUITA esta línea porque ya no necesitas pasar total_incidencia global
    return render(request, "construccion1app/dashboard_proyecto.html", {
        "proyecto": proyecto,
        "niveles": datos_niveles,
        "avance_total": avance_total,
        "form": form,
        "mostrar_formulario": mostrar_formulario,
        # ❌ NO pases "total_incidencia" aquí
    })


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from .models import Proyecto,Nivel,Espacio,Actividad
from .forms import ProyectoForm,NivelForm,EspacioForm,ActividadForm,ModificarActividadForm

@login_required
def agregar_proyecto(request):
    # Validar tipo de usuario
    if request.user.tipo_usuario not in ["superadmin_empresa", "admin_empresa"]:
        raise PermissionDenied("No tienes permiso para crear proyectos.")

    if request.method == "POST":
        form = ProyectoForm(request.POST)
        if form.is_valid():
            proyecto = form.save(commit=False)
            proyecto.empresa = request.user.empresa  # Se asigna automáticamente
            proyecto.save()
            return redirect('mis_proyectos')  # Nombre de la vista donde listamos proyectos
    else:
        form = ProyectoForm()
    
    return render(request, "construccion1app/agregar_proyecto.html", {"form": form})


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponseForbidden
from .models import Proyecto
@login_required
def eliminar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)

    # Restringir acceso SOLO a superusuario y superadmin_empresa
    if not (request.user.tipo_usuario in ['superusuario', 'superadmin_empresa']):
        return HttpResponseForbidden("No tienes permisos para eliminar proyectos.")

    if request.method == "POST":
        proyecto.delete()
        messages.success(request, "✅ Proyecto eliminado correctamente")
        return redirect("mis_proyectos")  # Ajusta la redirección según tu app

    return redirect('dashboard_proyecto', proyecto_id=proyecto.id)



@login_required
def eliminar_nivel(request, pk):
    nivel = get_object_or_404(Nivel, pk=pk)

    if not (request.user.tipo_usuario in ['superusuario', 'superadmin_empresa']):
        return HttpResponseForbidden("No tienes permisos para eliminar el Nivel.")

    if request.method == "POST":
        nivel.delete()
        messages.success(request, "✅ Nivel eliminado correctamente")
        return redirect('dashboard_proyecto', proyecto_id=nivel.proyecto.id)

    return redirect('dashboard_proyecto', proyecto_id=nivel.proyecto.id)



@login_required
def agregar_nivel(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    # Validar tipo de usuario
    if request.user.tipo_usuario not in ["superadmin_empresa", "admin_empresa"]:
        raise PermissionDenied("No tienes permiso para crear niveles.")

    if request.method == "POST":
        form = NivelForm(request.POST, proyecto=proyecto)
        if form.is_valid():
            nivel = form.save(commit=False)
            nivel.proyecto = proyecto  # <-- asignar al proyecto correcto
            nivel.save()
            return redirect('dashboard_proyecto', proyecto_id=proyecto.id)
        else:
            # Si el formulario no es válido, volver a renderizar la plantilla con errores
            return render(request, "construccion1app/dashboard_proyecto.html", {"form": form, "proyecto": proyecto})
    else:
        form = NivelForm(proyecto=proyecto)
        return render(request, "construccion1app/dashboard_proyecto.html", {"form": form, "proyecto": proyecto})
    


@login_required
def sumar_incidencias(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)
    suma = Nivel.objects.filter(proyecto=proyecto).aggregate(total=Sum("incidencia"))["total"] or 0
    return JsonResponse({"suma": suma})



from django.db.models import Sum

@login_required
def agregar_espacio(request, nivel_id):
    nivel = get_object_or_404(Nivel, id=nivel_id)
    total_actual = nivel.espacios.aggregate(total=Sum('incidencia'))['total'] or 0


    if request.method == "POST":
        form = EspacioForm(request.POST, nivel=nivel)
        if form.is_valid():
            espacio = form.save(commit=False)
            espacio.nivel = nivel
            espacio.save()
            messages.success(request, "Espacio agregado con éxito.")
            return redirect('dashboard_proyecto', proyecto_id=espacio.nivel.proyecto.id)
    else:
        form = EspacioForm(nivel=nivel)

    return render(request, 'construccion1app/agregar_espacio.html', {
        'form': form,
        'nivel': nivel,
        'total_actual': total_actual,  # 👉 aquí pasamos el total al template
    })

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponseForbidden
from .models import Espacio
@login_required
def eliminar_espacio(request, pk):
    espacio = get_object_or_404(Espacio, pk=pk)

    if not (request.user.is_superuser or request.user.tipo_usuario == 'superadmin_empresa'):
        return HttpResponseForbidden("No tienes permisos para eliminar espacios.")

    if request.method == "POST":
        proyecto_id = espacio.nivel.proyecto.id  # para redirigir al detalle del proyecto
        espacio.delete()
        messages.success(request, "✅ Espacio eliminado correctamente")
        return redirect('dashboard_proyecto', proyecto_id=espacio.nivel.proyecto.id)

    return redirect('dashboard_proyecto', proyecto_id=espacio.nivel.proyecto.id)


@login_required 
def agregar_actividad(request, espacio_id):     
    espacio = get_object_or_404(Espacio, id=espacio_id)      
    
    # Validar tipo de usuario     
    if request.user.tipo_usuario not in ['superadmin_empresa', 'admin_empresa', 'superuser']:         
        raise PermissionDenied("No tienes permiso para crear una actividad")      
    
    if request.method == "POST":         
        form = ActividadForm(
            request.POST, 
            request.FILES,  # Agregar files para manejar uploads
            user=request.user, 
            espacio=espacio  # Pasar espacio como parámetro
        )
        
        if form.is_valid():             
            # Simplificar el guardado - dejar que el form.save() maneje todo
            actividad = form.save()
            
            # Obtener el nombre del departamento correctamente
            espacio_nombre = espacio.nombre  # El espacio ES el departamento
            
            mensaje_base = f'Actividad agregada: "{actividad.nombre}" del espacio "{espacio_nombre}"'
            
            # Agregar información sobre la relación si existe
            if hasattr(form, 'relacion_establecida') and form.relacion_establecida:
                messages.success(request, f'{mensaje_base}. {form.relacion_establecida}')
            else:
                messages.success(request, mensaje_base)
                
            return redirect('dashboard_proyecto', proyecto_id=espacio.nivel.proyecto.id)
        
        else:
            messages.error(request, 'Por favor corrige los errores del formulario')
            
    else:         
        form = ActividadForm(
            user=request.user, 
            espacio=espacio  # Pasar espacio como parámetro
        )      
    
    return render(request, 'construccion1app/agregar_actividad.html', {
        'form': form, 
        'espacio': espacio,
        'user': request.user,  # Pasar user al template
    })


@login_required
def mapa_actividades(request, espacio_id):
    espacio = get_object_or_404(Espacio, id=espacio_id)
    actividades = Actividad.objects.filter(espacio=espacio)

    # Construir lista de nodos y relaciones
    nodes = []
    edges = []

    for act in actividades:
        nodes.append({
            'id': act.id,
            'label': act.nombre,
            'estado': act.estado_ejecucion,
            'inicio': act.fecha_inicio.isoformat() if act.fecha_inicio else None
        })
        if act.predecesora:
            edges.append({
                'from': act.predecesora.id,
                'to': act.id
            })
        if act.sucesora:
            edges.append({
                'from': act.id,
                'to': act.sucesora.id
            })

    return JsonResponse({'nodes': nodes, 'edges': edges})


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponseForbidden
from django.urls import reverse
@login_required
def modificar_actividad(request, actividad_id):

    # Obtener la actividad
    try:
        actividad = Actividad.objects.get(id=actividad_id)
    except Actividad.DoesNotExist:
        messages.error(request, "❌ La actividad que intentas modificar ya no existe.")
        return redirect('dashboard_proyecto', proyecto_id=actividad.espacio.nivel.proyecto.id)
    sucesora_habilitada = None
    
    # Verificar permisos (opcional)
    if not (
        request.user.tipo_usuario in ['superadmin_empresa', 'admin_empresa', 'superuser','calidad']
        or request.user == actividad.asignado  # 👈 Permitir al usuario asignado
    ):
        raise PermissionDenied("No tienes permiso para modificar esta actividad")
    
    if request.method == 'POST':
        # Crear formulario con datos POST y usuario
        form = ModificarActividadForm(
            data=request.POST, 
            files=request.FILES,
            instance=actividad,                     
            user=request.user
        )
        
        if form.is_valid():
            actividad_actualizada = form.save()
            
            # Obtener el nombre de la sucesora habilitada si existe
            if hasattr(form, 'sucesora_habilitada_nombre') and form.sucesora_habilitada_nombre:
                sucesora_habilitada = form.sucesora_habilitada_nombre
            
            # Mostrar mensaje apropiado
            if sucesora_habilitada:
                messages.success(
                    request, 
                    f'Actividad actualizada correctamente. La actividad sucesora "{sucesora_habilitada}" ha sido habilitada.'
                )
            else:
                messages.success(request, 'Actividad actualizada correctamente')
                
            return redirect( f"{reverse('dashboard_proyecto', args=[actividad.espacio.nivel.proyecto.id])}#actividad-{actividad.id}")
        else:
            messages.error(request, 'Por favor corrige los errores del formulario')
    
    else:
        # Crear formulario vacío con usuario
        form = ModificarActividadForm(
            instance=actividad,
            user=request.user
        )
    
    return render(request, 'construccion1app/modificar_actividad.html', {
        'form': form,
        'actividad': actividad,
        'user': request.user,
        'espacio': actividad.espacio,
        'sucesora_habilitada': sucesora_habilitada,  # También pasamos el user al template
    })


from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponseForbidden
from .models import Actividad
@login_required
def eliminar_actividad(request, pk):
    actividad = get_object_or_404(Actividad, pk=pk)

    # Control de permisos
    if not (
        request.user.is_superuser
        or request.user.tipo_usuario in ["superadmin_empresa", "admin_empresa"]
       
    ):
        return HttpResponseForbidden("No tienes permisos para eliminar esta actividad.")

    if request.method == "POST":
        proyecto_id = actividad.espacio.nivel.proyecto.id

        actividad.delete()
        messages.success(request, "✅ Actividad eliminada correctamente")
        return redirect('dashboard_proyecto', proyecto_id=actividad.espacio.nivel.proyecto.id)


    return redirect('dashboard_proyecto', proyecto_id=actividad.espacio.nivel.proyecto.id)

# context_processors.py
from .models import Notificacion
@login_required
def notificaciones_context(request):
    if request.user.is_authenticated:
        notificaciones_no_leidas = request.user.notificaciones.filter(leida=False).count()
    else:
        notificaciones_no_leidas = 0
    return {"notificaciones_no_leidas": notificaciones_no_leidas}

@login_required
def detalle_notificacion(request, pk):
    # Obtener la notificación del usuario
    notif = get_object_or_404(Notificacion, pk=pk, usuario=request.user)
    
    # Marcar como leída
    notif.leida = True
    notif.save()
    
    # Redirigir a la edición de la actividad asociada
    if notif.actividad:  # Asegúrate de que la notificación tenga FK a Actividad
        return redirect('modificar_actividad', actividad_id=notif.actividad.pk)
    
    return redirect('lista_notificaciones')  # O a otra página si no hay actividad asociada



from .models import Notificacion
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render, redirect

@login_required
def lista_notificaciones(request):
    # Solo del usuario logueado
    notificaciones_list = request.user.notificaciones.all().order_by('-fecha_creacion')
    
    # Paginación: 30 notificaciones por página
    paginator = Paginator(notificaciones_list, 30)
    page_number = request.GET.get('page', 1)
    notificaciones = paginator.get_page(page_number)
    
    return render(request, 'construccion1app/lista_notificaciones.html', {
        'notificaciones': notificaciones
    })

@login_required
def marcar_todas_leidas(request):
    """Marca todas las notificaciones del usuario como leídas"""
    request.user.notificaciones.filter(leida=False).update(leida=True)
    return redirect('lista_notificaciones')






from django.db.models import Sum
from .models import Actividad
@login_required
def incidencia_restante(request, espacio_id):
    exclude_id = request.GET.get("exclude_id")
    qs = Actividad.objects.filter(espacio_id=espacio_id)
    if exclude_id:
        qs = qs.exclude(id=exclude_id)

    total_otras = qs.aggregate(total=Sum("incidencia"))["total"] or 0
    return JsonResponse({"total_otras": total_otras})



import pandas as pd
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import Actividad, Espacio
from decimal import Decimal
@login_required
def importar_actividades(request, espacio_id):
    espacio = get_object_or_404(Espacio, id=espacio_id)

    if request.method == "POST" and request.FILES.get("archivo_excel"):
        archivo = request.FILES["archivo_excel"]

        try:
            df = pd.read_excel(archivo)

            # Campos obligatorios
            campos_requeridos = [
                "nombre", "avance", "incidencia",
                "estado_ejecucion", "estado_asignacion", "habilitada"
            ]
            
            # Validar que existan todas las columnas
            for campo in campos_requeridos:
                if campo not in df.columns:
                    raise ValueError(f"❌ Falta la columna obligatoria: '{campo}'")

            # Validar que no haya valores nulos o vacíos
            for campo in campos_requeridos:
                if df[campo].isnull().any():
                    filas_nulas = df[df[campo].isnull()].index.tolist()
                    raise ValueError(
                        f"❌ La columna '{campo}' tiene valores vacíos en las filas: {[i+2 for i in filas_nulas]}"
                    )
                if (df[campo].astype(str).str.strip() == "").any():
                    filas_vacias = df[df[campo].astype(str).str.strip() == ""].index.tolist()
                    raise ValueError(
                        f"❌ La columna '{campo}' tiene valores vacíos en las filas: {[i+2 for i in filas_vacias]}"
                    )

            # Validar valores permitidos para estado_ejecucion
            estados_ejecucion_validos = [
                "no_ejecutada", "ejecucion", "ejecutada", "revisada", "observada"
            ]
            estados_invalidos_ejecucion = df[
                ~df["estado_ejecucion"].isin(estados_ejecucion_validos)
            ]
            if not estados_invalidos_ejecucion.empty:
                valores_invalidos = estados_invalidos_ejecucion["estado_ejecucion"].unique().tolist()
                filas_invalidas = estados_invalidos_ejecucion.index.tolist()
                raise ValueError(
                    f"❌ Valores inválidos en 'estado_ejecucion': {valores_invalidos} "
                    f"en las filas: {[i+2 for i in filas_invalidas]}. "
                    f"Valores permitidos: {estados_ejecucion_validos}"
                )

            # Validar valores permitidos para estado_asignacion
            estados_asignacion_validos = ["POR_ASIGNAR", "ASIGNADA", "NO_ASIGNADA"]
            estados_invalidos_asignacion = df[
                ~df["estado_asignacion"].isin(estados_asignacion_validos)
            ]
            if not estados_invalidos_asignacion.empty:
                valores_invalidos = estados_invalidos_asignacion["estado_asignacion"].unique().tolist()
                filas_invalidas = estados_invalidos_asignacion.index.tolist()
                raise ValueError(
                    f"❌ Valores inválidos en 'estado_asignacion': {valores_invalidos} "
                    f"en las filas: {[i+2 for i in filas_invalidas]}. "
                    f"Valores permitidos: {estados_asignacion_validos}"
                )

            # Validar que 'habilitada' sea booleano (True/False, 1/0, Si/No)
            valores_habilitada_validos = [True, False, 1, 0, "True", "False", "1", "0", 
                                          "true", "false", "Si", "No", "SI", "NO", 
                                          "si", "no", "Sí", "sí", "SÍ"]
            habilitada_invalida = df[
                ~df["habilitada"].isin(valores_habilitada_validos)
            ]
            if not habilitada_invalida.empty:
                valores_invalidos = habilitada_invalida["habilitada"].unique().tolist()
                filas_invalidas = habilitada_invalida.index.tolist()
                raise ValueError(
                    f"❌ Valores inválidos en 'habilitada': {valores_invalidos} "
                    f"en las filas: {[i+2 for i in filas_invalidas]}. "
                    f"Valores permitidos: True/False, 1/0, Si/No"
                )

            # Validar que 'avance' sea numérico entre 0 y 100
            if not pd.api.types.is_numeric_dtype(df["avance"]):
                raise ValueError("❌ La columna 'avance' debe contener solo valores numéricos")
            
            avances_invalidos = df[(df["avance"] < 0) | (df["avance"] > 100)]
            if not avances_invalidos.empty:
                filas_invalidas = avances_invalidos.index.tolist()
                valores_invalidos = avances_invalidos["avance"].tolist()
                raise ValueError(
                    f"❌ La columna 'avance' debe estar entre 0 y 100. "
                    f"Valores inválidos: {valores_invalidos} en las filas: {[i+2 for i in filas_invalidas]}"
                )

            # Validar que 'incidencia' sea numérica y positiva
            if not pd.api.types.is_numeric_dtype(df["incidencia"]):
                raise ValueError("❌ La columna 'incidencia' debe contener solo valores numéricos")
            
            incidencias_invalidas = df[df["incidencia"] <= 0]
            if not incidencias_invalidas.empty:
                filas_invalidas = incidencias_invalidas.index.tolist()
                raise ValueError(
                    f"❌ La columna 'incidencia' debe tener valores mayores a 0. "
                    f"Filas con valores inválidos: {[i+2 for i in filas_invalidas]}"
                )
                        # Validar que 'avance' sea numérico
            if not pd.api.types.is_numeric_dtype(df["avance"]):
                raise ValueError("❌ La columna 'avance' debe contener solo valores numéricos")

            # Validar que 'avance' sea exactamente 0 o 100 (sin decimales)
            avances_invalidos = df[~df["avance"].isin([0, 0.0, 100, 100.0])]
            if not avances_invalidos.empty:
                filas_invalidas = avances_invalidos.index.tolist()
                valores_invalidos = avances_invalidos["avance"].tolist()
                raise ValueError(
                    f"❌ La columna 'avance' solo puede contener los valores 0 o 100. "
                    f"Valores inválidos: {valores_invalidos} en las filas: {[i+2 for i in filas_invalidas]}"
    )

            # Validar suma de incidencias
            suma_incidencia = df["incidencia"].sum()
            if round(float(suma_incidencia), 2) != 100.00:
                raise ValueError(
                    f"❌ La suma de incidencias debe ser exactamente 100.00. "
                    f"La suma actual es: {round(suma_incidencia, 2)}"
                )

            # Validar que no haya nombres duplicados
            nombres_duplicados = df[df.duplicated(subset=["nombre"], keep=False)]
            if not nombres_duplicados.empty:
                nombres_dup = nombres_duplicados["nombre"].unique().tolist()
                raise ValueError(
                    f"❌ Existen nombres de actividades duplicados: {nombres_dup}"
                )

            # Normalizar valores booleanos para 'habilitada'
            def normalizar_booleano(valor):
                if valor in [True, 1, "True", "1", "true", "Si", "SI", "si", "Sí", "sí", "SÍ"]:
                    return True
                return False

            # Crear actividades dentro de una transacción
            from django.db import transaction
            
            with transaction.atomic():
                actividades_creadas = 0
                for index, row in df.iterrows():
                    Actividad.objects.create(
                        espacio=espacio,
                        nombre=str(row["nombre"]).strip(),
                        avance=float(row["avance"]),
                        incidencia=Decimal(str(row["incidencia"])),
                        estado_ejecucion=str(row["estado_ejecucion"]).strip(),
                        estado_asignacion=str(row["estado_asignacion"]).strip(),
                        habilitada=normalizar_booleano(row["habilitada"]),
                    )
                    actividades_creadas += 1

            messages.success(
                request, 
                f"✅ {actividades_creadas} actividades importadas con éxito al espacio '{espacio.nombre}'"
            )

        except ValueError as ve:
            # Errores de validación controlados
            messages.error(request, str(ve))

        except pd.errors.EmptyDataError:
            messages.error(request, "❌ El archivo Excel está vacío")

        except pd.errors.ParserError:
            messages.error(request, "❌ El archivo no es un Excel válido o está corrupto")

        except KeyError as ke:
            messages.error(request, f"❌ Columna no encontrada en el archivo: {str(ke)}")

        except Exception as e:
            # Errores inesperados
            messages.error(request, f"⚠️ Error inesperado al procesar el archivo: {str(e)}")

        return redirect("dashboard_proyecto", proyecto_id=espacio.nivel.proyecto.id)

    return render(request, "construccion1app/importar_actividades.html", {"espacio": espacio})


from django.http import HttpResponse
import pandas as pd
from io import BytesIO
@login_required
def descargar_plantilla_actividades(request, espacio_id):
    """Genera y descarga una plantilla Excel para importar actividades"""
    espacio = get_object_or_404(Espacio, id=espacio_id)
    
    # Crear DataFrame con las columnas requeridas y datos de ejemplo
    datos_ejemplo = {
        'nombre': ['Actividad Ejemplo 1', 'Actividad Ejemplo 2'],
        'avance': [0, 0],
        'incidencia': [50.00, 50.00],
        'estado_ejecucion': ['no_ejecutada', 'no_ejecutada'],
        'estado_asignacion': ['no_asignada', 'asignada'],
        'habilitada': [True, True]
    }
    
    df = pd.DataFrame(datos_ejemplo)
    
    # Crear archivo Excel en memoria
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Actividades')
        
        # Opcional: ajustar ancho de columnas
        worksheet = writer.sheets['Actividades']
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Preparar respuesta HTTP
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="plantilla_actividades_{espacio.nombre}.xlsx"'
    
    return response

import datetime
from datetime import date
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.db.models import Q
from .models import Proyecto, Actividad

@login_required
def programacion_obra(request, proyecto_id):
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    # Validar tipo de usuario
    if request.user.tipo_usuario not in ["superadmin_empresa", "admin_empresa","calidad","supervisor_constructor","superuser"]:
        raise PermissionDenied("No tienes permiso para acceder a la programación de obra.")

    vista = request.GET.get("vista", "kanban")

    # === Capturar filtros GET ===
    nombre = request.GET.get("nombre", "")
    asignado_id = request.GET.get("asignado", "")
    estado_ejecucion = request.GET.get("estado_ejecucion", "")
    estado_asignacion = request.GET.get("estado_asignacion", "")
    habilitada = request.GET.get("habilitada", "")
    fecha_inicio = request.GET.get("fecha_inicio", "")
    fecha_fin = request.GET.get("fecha_fin", "")
    avance_min = request.GET.get("avance_min", "")
    avance_max = request.GET.get("avance_max", "")
    nivel_id = request.GET.get("nivel", "")
    espacio_nombre = request.GET.get("espacio")

    # === Filtrar actividades ===
    actividades_qs = Actividad.objects.filter(espacio__nivel__proyecto=proyecto)

    if nombre:
        actividades_qs = actividades_qs.filter(nombre__icontains=nombre)
    if asignado_id:
        actividades_qs = actividades_qs.filter(asignado_id=asignado_id)
    if estado_ejecucion:
        actividades_qs = actividades_qs.filter(estado_ejecucion=estado_ejecucion)
    if estado_asignacion:
        actividades_qs = actividades_qs.filter(estado_asignacion=estado_asignacion)
    if habilitada:
        actividades_qs = actividades_qs.filter(habilitada=(habilitada == "true"))
    if fecha_inicio:
        actividades_qs = actividades_qs.filter(fecha_inicio__gte=fecha_inicio)
    if fecha_fin:
        actividades_qs = actividades_qs.filter(fecha_fin__lte=fecha_fin)
    if avance_min:
        actividades_qs = actividades_qs.filter(avance__gte=avance_min)
    if avance_max:
        actividades_qs = actividades_qs.filter(avance__lte=avance_max)
    if nivel_id:
        actividades_qs = actividades_qs.filter(espacio__nivel_id=nivel_id)
    if espacio_nombre:
        actividades_qs = actividades_qs.filter(espacio__nombre__icontains=espacio_nombre)

    # === Estados ===
    ESTADO_EJECUCION_CHOICES = {
        "no_ejecutada": "No ejecutada",
        "ejecucion": "En ejecución",
        "ejecutada": "Ejecutada",
        "observada": "Observada",
        "revisada": "Revisada",
    }

    estados = {key: [] for key in ESTADO_EJECUCION_CHOICES.keys()}

    # Solo procesar si hay actividades
    for actividad in actividades_qs.select_related("espacio__nivel", "asignado"):
        # Validación de seguridad por si acaso
        if not actividad.espacio or not actividad.espacio.nivel:
            continue
            
        nivel = actividad.espacio.nivel
        espacio = actividad.espacio
        inicio = actividad.fecha_inicio
        fin = actividad.fecha_fin
        dias = (fin - inicio).days + 1 if inicio and fin else None

        estados[actividad.estado_ejecucion].append({
            "nivel": nivel.nombre,
            "espacio": espacio.nombre,
            "actividad": actividad,
            "asignado": actividad.asignado,
            "dias": dias,
        })

    context = {
        "proyecto": proyecto,
        "estados": estados,
        "ESTADO_EJECUCION_CHOICES": ESTADO_EJECUCION_CHOICES,
        "vista": vista,
        "hoy": date.today(),
        "asignados": Actividad.objects.filter(
            espacio__nivel__proyecto=proyecto,
            asignado__isnull=False
        ).values_list("asignado__id", "asignado__username").distinct(),
        "filtros": {
            "nombre": nombre,
            "asignado": asignado_id,
            "estado_ejecucion": estado_ejecucion,
            "estado_asignacion": estado_asignacion,
            "habilitada": habilitada,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "avance_min": avance_min,
            "avance_max": avance_max,
        },
    }

    return render(request, "construccion1app/programacion.html", context)
    
    


from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from decimal import Decimal

@login_required
def exportar_actividades_excel(request, proyecto_id):
    """
    Exporta actividades a Excel con formato optimizado para Power BI
    Respeta todos los filtros aplicados en la vista
    """
    proyecto = get_object_or_404(Proyecto, id=proyecto_id)

    # === Capturar TODOS los filtros (igual que en programacion_obra) ===
    nombre = request.GET.get("nombre", "")
    asignado_id = request.GET.get("asignado", "")
    estado_ejecucion = request.GET.get("estado_ejecucion", "")
    estado_asignacion = request.GET.get("estado_asignacion", "")
    habilitada = request.GET.get("habilitada", "")
    fecha_inicio = request.GET.get("fecha_inicio", "")
    fecha_fin = request.GET.get("fecha_fin", "")
    avance_min = request.GET.get("avance_min", "")
    avance_max = request.GET.get("avance_max", "")
    nivel_id = request.GET.get("nivel", "")
    espacio_nombre = request.GET.get("espacio", "")

    # === Aplicar EXACTAMENTE los mismos filtros ===
    actividades_qs = Actividad.objects.filter(
        espacio__nivel__proyecto=proyecto
    ).select_related(
        'espacio__nivel',
        'asignado',
        'predecesora',
        'sucesora'
    )

    if nombre:
        actividades_qs = actividades_qs.filter(nombre__icontains=nombre)
    if asignado_id:
        actividades_qs = actividades_qs.filter(asignado_id=asignado_id)
    if estado_ejecucion:
        actividades_qs = actividades_qs.filter(estado_ejecucion=estado_ejecucion)
    if estado_asignacion:
        actividades_qs = actividades_qs.filter(estado_asignacion=estado_asignacion)
    if habilitada:
        actividades_qs = actividades_qs.filter(habilitada=(habilitada == "true"))
    if fecha_inicio:
        actividades_qs = actividades_qs.filter(fecha_inicio__gte=fecha_inicio)
    if fecha_fin:
        actividades_qs = actividades_qs.filter(fecha_fin__lte=fecha_fin)
    if avance_min:
        actividades_qs = actividades_qs.filter(avance__gte=avance_min)
    if avance_max:
        actividades_qs = actividades_qs.filter(avance__lte=avance_max)
    if nivel_id:
        actividades_qs = actividades_qs.filter(espacio__nivel_id=nivel_id)
    if espacio_nombre:
        actividades_qs = actividades_qs.filter(espacio__nombre__icontains=espacio_nombre)

    # === Crear archivo Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Actividades"

    # === Estilos para el encabezado ===
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === Definir columnas optimizadas para Power BI ===
    headers = [
        "ID_Actividad",
        "Proyecto",
        "Nivel",
        "Espacio",
        "Actividad",
        "Estado_Ejecucion",
        "Estado_Asignacion",
        "Avance_%",
        "Incidencia_%",
        "Avance_Ponderado",
        "Asignado_A",
        "Email_Asignado",
        "Fecha_Inicio",
        "Fecha_Fin",
        "Duracion_Dias",
        "Habilitada",
        "Aprobacion_Calidad",
        "Tiene_Foto",
        "Tiene_Justificacion",
        "Predecesora",
        "Sucesora",
        "Fecha_Exportacion"
    ]

    # === Escribir encabezados ===
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # === Escribir datos ===
    row_num = 2
    for actividad in actividades_qs:
        # Calcular duración
        duracion = None
        if actividad.fecha_inicio and actividad.fecha_fin:
            duracion = (actividad.fecha_fin - actividad.fecha_inicio).days + 1

        # Avance ponderado
        avance_ponderado = float(actividad.calcular_avance_ponderado())

        # Datos del asignado
        asignado_nombre = actividad.asignado.username if actividad.asignado else ""
        asignado_email = actividad.asignado.email if actividad.asignado else ""

        # Fila de datos
        row_data = [
            actividad.id,
            proyecto.nombre,
            actividad.espacio.nivel.nombre,
            actividad.espacio.nombre,
            actividad.nombre,
            actividad.get_estado_ejecucion_display(),
            actividad.get_estado_asignacion_display(),
            float(actividad.avance),
            float(actividad.incidencia),
            avance_ponderado,
            asignado_nombre,
            asignado_email,
            actividad.fecha_inicio,
            actividad.fecha_fin,
            duracion,
            "Sí" if actividad.habilitada else "No",
            "Sí" if actividad.aprobacion_calidad else "No",

            "Sí" if actividad.justificacion else "No",
            actividad.predecesora.nombre if actividad.predecesora else "",
            actividad.sucesora.nombre if actividad.sucesora else "",
            date.today()
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            
            # Alineación según tipo de dato
            if isinstance(value, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(value, date):
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = 'DD/MM/YYYY'
            else:
                cell.alignment = Alignment(horizontal="left")

        row_num += 1

    # === Ajustar anchos de columna ===
    column_widths = {
        'A': 12,  # ID
        'B': 20,  # Proyecto
        'C': 15,  # Nivel
        'D': 15,  # Espacio
        'E': 30,  # Actividad
        'F': 18,  # Estado Ejecución
        'G': 18,  # Estado Asignación
        'H': 10,  # Avance
        'I': 12,  # Incidencia
        'J': 15,  # Avance Ponderado
        'K': 20,  # Asignado
        'L': 25,  # Email
        'M': 13,  # Fecha Inicio
        'N': 13,  # Fecha Fin
        'O': 12,  # Duración
        'P': 11,  # Habilitada
        'Q': 16,  # Aprobación Calidad
        'S': 18,  # Tiene Justificación
        'T': 20,  # Predecesora
        'U': 20,  # Sucesora
        'V': 16,  # Fecha Exportación
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # === Congelar primera fila ===
    ws.freeze_panes = 'A2'

    # === Preparar respuesta HTTP ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nombre del archivo con fecha
    filename = f"Actividades_{proyecto.nombre}_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Guardar archivo
    wb.save(response)
    
    return response


from django.contrib.auth.decorators import login_required

from django.views.decorators.http import require_POST


# views.py




# views.py
from django.shortcuts import render
from django.conf import settings
from django.contrib.auth.decorators import login_required

@login_required
def base_view(request):
    """
    Renderiza la base y pasa la clave VAPID al template
    """
    context = {
        'VAPID_PUBLIC_KEY': settings.WEBPUSH_SETTINGS['VAPID_PUBLIC_KEY']
    }
    return render(request, 'construccion1app/base.html', context)


@login_required
@csrf_exempt
def save_fcm_token(request):
    """
    Guarda el token FCM del usuario en la base de datos
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            token = data.get('token')
            
            print(f"🔍 Usuario actual: {request.user.email}")
            print(f"🔍 Token anterior: {request.user.fcm_token[:50] if request.user.fcm_token else 'NINGUNO'}...")
            print(f"🔍 Token nuevo: {token[:50]}..." if token else "❌ Token vacío")
            
            if token:
                # VERIFICAR: Si el token es el mismo, no hacer nada
                if request.user.fcm_token == token:
                    print(f"ℹ️ Token sin cambios para {request.user.email}")
                    return JsonResponse({
                        'status': 'success', 
                        'message': 'Token ya estaba guardado',
                        'user': request.user.email
                    })
                
                # REEMPLAZAR el token (no concatenar)
                request.user.fcm_token = token
                request.user.save(update_fields=['fcm_token'])
                
                # Verificar que se guardó correctamente
                request.user.refresh_from_db()
                token_guardado = request.user.fcm_token
                
                print(f"✅ Token REEMPLAZADO para {request.user.email}")
                print(f"✅ Token guardado: {token_guardado[:50]}...")
                
                # Validar que el token guardado es correcto
                if token_guardado != token:
                    print(f"⚠️ ADVERTENCIA: Token guardado no coincide!")
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Error al guardar token'
                    }, status=500)
                
                return JsonResponse({
                    'status': 'success', 
                    'message': 'Token guardado correctamente',
                    'user': request.user.email,
                    'token_length': len(token)
                })
            else:
                print("❌ Token no proporcionado en la solicitud")
                return JsonResponse({
                    'status': 'error', 
                    'message': 'Token no proporcionado'
                }, status=400)
                
        except Exception as e:
            print(f"❌ Error al guardar token FCM: {e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error', 
                'message': str(e)
            }, status=500)
    
    return JsonResponse({
        'status': 'error', 
        'message': 'Método no permitido'
    }, status=405)





from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
import firebase_admin
from firebase_admin import messaging

@login_required
@require_http_methods(["GET"])
def test_notification(request):
    """Vista temporal para probar notificaciones push"""
    try:
        # 1. Verificar que Firebase está inicializado
        if not firebase_admin._apps:
            return JsonResponse({
                'status': 'error',
                'message': 'Firebase no está inicializado. Verifica firebase-key.json'
            }, status=500)
        
        # 2. Verificar usuario autenticado
        if not request.user.is_authenticated:
            return JsonResponse({
                'status': 'error',
                'message': 'Usuario no autenticado'
            }, status=401)
        
        # 3. Verificar token existe
        token = getattr(request.user, 'fcm_token', None)
        
        print(f"🔍 Usuario: {request.user.email}")
        print(f"🔍 Tiene atributo fcm_token: {hasattr(request.user, 'fcm_token')}")
        print(f"🔍 Token: {token[:50] if token else 'NINGUNO'}...")
        
        if not token:
            return JsonResponse({
                'status': 'error',
                'message': 'No hay token FCM guardado para este usuario',
                'suggestion': 'Recarga la página para generar un nuevo token',
                'user': request.user.email
            }, status=400)
        
        # 4. Verificar que el token no esté vacío o corrupto
        if len(token) < 100:  # Los tokens FCM son largos
            return JsonResponse({
                'status': 'error',
                'message': 'Token FCM parece inválido o corrupto',
                'token_length': len(token),
                'suggestion': 'Borra el token y recarga la página'
            }, status=400)
        
        # 5. Intentar enviar notificación
        print(f"📤 Intentando enviar notificación...")
        
        message = messaging.Message(
            notification=messaging.Notification(
                title="🧪 Prueba de Notificación",
                body=f"Hola {request.user.first_name or request.user.email}, las notificaciones funcionan!"
            ),
            token=token,
            webpush=messaging.WebpushConfig(
                notification=messaging.WebpushNotification(
                    icon="/static/construccion1app/img/logo2.jpeg",
                    badge="/static/construccion1app/img/logo2.jpeg",
                    vibrate=[200, 100, 200]
                )
            )
        )
        
        response = messaging.send(message)
        
        print(f"✅ Notificación enviada exitosamente")
        print(f"✅ Response ID: {response}")
        
        return JsonResponse({
            'status': 'success',
            'message': 'Notificación enviada correctamente',
            'response_id': response,
            'user': request.user.email
        })
        
    except messaging.UnregisteredError as e:
        print(f"❌ Token no registrado o expirado: {e}")
        return JsonResponse({
            'status': 'error',
            'message': 'Token FCM inválido o expirado',
            'suggestion': 'Recarga la página para generar un nuevo token',
            'error_type': 'UnregisteredError'
        }, status=400)
        
    except messaging.InvalidArgumentError as e:
        print(f"❌ Argumento inválido: {e}")
        return JsonResponse({
            'status': 'error',
            'message': 'Error en los datos de la notificación',
            'error': str(e),
            'error_type': 'InvalidArgumentError'
        }, status=400)
        
    except Exception as e:
        print(f"❌ Error inesperado:")
        import traceback
        traceback.print_exc()
        
        return JsonResponse({
            'status': 'error',
            'message': f'Error al enviar notificación: {str(e)}',
            'error_type': type(e).__name__,
            'suggestion': 'Revisa los logs del servidor'
        }, status=500)
    
@login_required
def diagnostico_firebase(request):
    """Vista para diagnosticar problemas de Firebase"""
    import os
    from django.conf import settings
    
    diagnostico = {
        'firebase_initialized': len(firebase_admin._apps) > 0,
        'credentials_path': os.path.join(settings.BASE_DIR, 'credentials', 'firebase-key.json'),
        'credentials_exist': os.path.exists(os.path.join(settings.BASE_DIR, 'credentials', 'firebase-key.json')),
        'user_authenticated': request.user.is_authenticated,
        'user_email': request.user.email if request.user.is_authenticated else None,
        'has_fcm_token_field': hasattr(request.user, 'fcm_token'),
        'fcm_token_length': len(request.user.fcm_token) if hasattr(request.user, 'fcm_token') and request.user.fcm_token else 0,
        'fcm_token_preview': request.user.fcm_token[:50] if hasattr(request.user, 'fcm_token') and request.user.fcm_token else 'NINGUNO'
    }
    
    return JsonResponse(diagnostico, json_dumps_params={'indent': 2})